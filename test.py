import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Budget"

# Set the headers for the columns
headers = ["Category", "Planned", "Actual", "Difference"]
categories = [
    ("Income", 3000, 3000, 0),
    ("Rent", 1000, 1000, 0),
    ("Groceries", 300, 350, -50),
    ("Utilities", 150, 120, 30),
    ("Transportation", 100, 120, -20),
    ("Entertainment", 200, 180, 20),
    ("Savings", 250, 200, -50),
]

# Apply headers with bold font, background color, and center alignment
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    # Add border around the header cells
    cell.border = Border(
        top=Side(style="thin"), 
        bottom=Side(style="thin"), 
        left=Side(style="thin"), 
        right=Side(style="thin")
    )

# Add categories and data with alternating row colors and borders
row_fill_odd = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
row_fill_even = PatternFill(start_color="EAF1FB", end_color="EAF1FB", fill_type="solid")
font = Font(name="Calibri", size=12)
alignment = Alignment(horizontal="left", vertical="center")

for row_num, (category, planned, actual, diff) in enumerate(categories, 2):
    fill = row_fill_odd if row_num % 2 == 0 else row_fill_even
    ws.cell(row=row_num, column=1, value=category).fill = fill
    ws.cell(row=row_num, column=2, value=planned).fill = fill
    ws.cell(row=row_num, column=3, value=actual).fill = fill
    ws.cell(row=row_num, column=4, value=diff).fill = fill
    
    # Apply font and alignment
    for col_num in range(1, 5):
        cell = ws.cell(row=row_num, column=col_num)
        cell.font = font
        cell.alignment = alignment
        # Add borders around the data cells
        cell.border = Border(
            top=Side(style="thin"), 
            bottom=Side(style="thin"), 
            left=Side(style="thin"), 
            right=Side(style="thin")
        )

# Adjust column width for better readability
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15

# Save the workbook
wb.save("styled_personal_budget.xlsx")

print("Styled Excel template created successfully!")
